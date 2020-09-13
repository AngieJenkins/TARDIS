from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pymorphy2
import requests
from bs4 import BeautifulSoup
import json

# Открываем файл, создаем частотный список

f = open('dom.txt', 'r', encoding = 'utf-8')
spisok = []

for string in f:
    words = string.split()
    for one in words:
        one = one.strip('0123456789#№%+=,.!?:;*—\/""-«»…”“() ')
        one = one.lower()
        spisok.append(one)

f.close()
frequency = Counter(spisok)

# Записываем частотный список в csv-файл, лемматизируем, ищем леммы с двумя "о"

csv = openpyxl.Workbook()
sheet = csv.active
morph = pymorphy2.MorphAnalyzer()

sheet.cell(row = 1, column = 1).value = 'словоформа'
sheet.cell(row = 1, column = 2).value = 'лемма'
sheet.cell(row = 1, column = 3).value = 'частота'
sheet.cell(row = 1, column = 4).value = 'две "о" в лемме'

m = 1
for word in frequency:
    m += 1
    sheet.cell(row = m, column = 1).value = frequency.most_common()[m-2][0]
    sheet.cell(row = m, column = 3).value = frequency.most_common()[m-2][1]
    
    cell = sheet.cell(row = m, column = 2)
    p = morph.parse(frequency.most_common()[m-2][0])
    collect = []
    for j in p:
        collect.append(str(j.normal_form))
    collect = set(collect)
    
    for lemma in collect:
        bukvy = list(lemma)
        bukvy_frq = Counter(bukvy)
        if bukvy_frq['о'] == 2:
            sheet.cell(row = m, column = 4).value = '+'
            
    cell.value = ' '.join(collect)

csv.save('Частотный_список.csv')

# Считываем текст с сайта, составляем частотный список

r = requests.post('http://lib.ru/POEZIQ/PESSOA/lirika.txt')
if r.status_code != 200:
    print('Error')

soup = BeautifulSoup(r.content, features = "html.parser")
text = soup.get_text()

slova = []
words = text.split()
for word in words:
    word = word.strip('0123456789#№%+=,.!?:;*—\/""-«»…”“() ')
    word = word.lower()
    slova.append(word)

frq_list = Counter(slova)
frq_json = open('frequency_list.json', 'w')
json.dump(frq_list, frq_json)
frq_json.close()
